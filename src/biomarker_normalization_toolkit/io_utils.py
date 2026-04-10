from __future__ import annotations

import csv
import json
import logging
from defusedxml.ElementTree import fromstring as _xml_fromstring
import xml.etree.ElementTree as ET
from pathlib import Path

logger = logging.getLogger(__name__)

from biomarker_normalization_toolkit.fhir import build_bundle
from biomarker_normalization_toolkit.models import NormalizationResult
from biomarker_normalization_toolkit.reporting import build_summary_report


REQUIRED_INPUT_COLUMNS = (
    "source_row_id",
    "source_test_name",
    "raw_value",
    "source_unit",
    "specimen_type",
    "source_reference_range",
)


def _find_duplicate_labels(labels: list[str]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for label in labels:
        if label in seen and label not in duplicates:
            duplicates.append(label)
        seen.add(label)
    return duplicates


def _extract_loinc_code(coding: list[dict]) -> str:
    for code in coding:
        if not isinstance(code, dict):
            continue
        if (code.get("system", "") or "").strip() == "http://loinc.org" and code.get("code"):
            return str(code["code"]).strip()
    return ""


def read_input(path: Path) -> list[dict[str, str]]:
    """Auto-detect format and read input file. Supports CSV, FHIR JSON, HL7v2, and C-CDA XML."""
    suffix = path.suffix.lower()
    if suffix == ".json":
        return read_fhir_input(path)
    if suffix in (".hl7", ".oru"):
        return read_hl7_input(path)
    if suffix == ".xml":
        return read_ccda_input(path)
    if suffix in (".xlsx", ".xls"):
        return read_excel_input(path)
    return read_input_csv(path)


def read_fhir_input(path: Path) -> list[dict[str, str]]:
    """Read a FHIR Bundle JSON and extract Observation resources into input rows."""
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(data, dict):
        raise ValueError("FHIR input must be a JSON object representing an Observation or Bundle.")

    if data.get("resourceType") == "Observation":
        entries = [{"resource": data}]
    elif data.get("resourceType") == "Bundle":
        entries = data.get("entry", [])
        if not isinstance(entries, list):
            raise ValueError("FHIR Bundle.entry must be an array.")
    else:
        raise ValueError(f"Unrecognized FHIR resourceType: {data.get('resourceType', 'none')}")

    def _index_specimen(resource: dict, index: dict[str, str], full_url: str = "") -> None:
        if not isinstance(resource, dict):
            return
        type_obj = resource.get("type", {})
        if not isinstance(type_obj, dict):
            return
        coding = type_obj.get("coding", [])
        if not isinstance(coding, list):
            coding = []
        specimen_display = (type_obj.get("text", "") or "").strip()
        if not specimen_display:
            for code in coding:
                if not isinstance(code, dict):
                    continue
                specimen_display = (code.get("display", "") or code.get("code", "") or "").strip()
                if specimen_display:
                    break
        if not specimen_display:
            return
        specimen_id = str(resource.get("id", "")).strip()
        references = [full_url]
        if specimen_id:
            references.extend((
                f"Specimen/{specimen_id}",
                specimen_id,
                f"#{specimen_id}",
            ))
            if full_url:
                references.append(f"{full_url}#{specimen_id}")
        for reference in references:
            if reference:
                index[reference] = specimen_display

    # Flatten: extract Observations from DiagnosticReport.contained as well.
    # Also index Specimen resources so Observation.specimen.reference can be resolved.
    all_observations: list[dict] = []
    specimen_display_by_reference: dict[str, str] = {}
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        resource = entry.get("resource", entry)
        if not isinstance(resource, dict):
            continue
        rt = resource.get("resourceType")
        full_url = str(entry.get("fullUrl", "")).strip()
        if rt == "Specimen":
            _index_specimen(resource, specimen_display_by_reference, full_url=full_url)
        elif rt == "Observation":
            all_observations.append(resource)
        contained_resources = resource.get("contained", [])
        if not isinstance(contained_resources, list):
            contained_resources = []
        for contained in contained_resources:
            if not isinstance(contained, dict):
                continue
            contained_type = contained.get("resourceType")
            if contained_type == "Observation":
                all_observations.append(contained)
            elif contained_type == "Specimen":
                _index_specimen(contained, specimen_display_by_reference, full_url=full_url)

    rows: list[dict[str, str]] = []
    for index, resource in enumerate(all_observations, start=1):

        code_obj = resource.get("code", {})
        if not isinstance(code_obj, dict):
            code_obj = {}
        coding = code_obj.get("coding", [])
        if not isinstance(coding, list):
            coding = []
        source_loinc = _extract_loinc_code(coding)
        test_name = (code_obj.get("text", "") or "").strip()
        if not test_name:
            for code in coding:
                if not isinstance(code, dict):
                    continue
                test_name = (code.get("display", "") or "").strip()
                if test_name:
                    break
        if not test_name:
            if source_loinc:
                test_name = source_loinc
        if not test_name:
            for code in coding:
                test_name = (code.get("code", "") or "").strip()
                if test_name:
                    break
        if not test_name:
            continue

        vq = resource.get("valueQuantity", {})
        if not isinstance(vq, dict):
            vq = {}
        value = vq.get("value")
        unit = vq.get("unit", vq.get("code", "")) if vq else ""

        # Fall back to other FHIR value types if no valueQuantity
        if value is None:
            if "valueString" in resource:
                value = resource["valueString"]
            elif "valueInteger" in resource:
                value = resource["valueInteger"]
            elif "valueCodeableConcept" in resource:
                cc = resource["valueCodeableConcept"]
                value = cc.get("text", "")
                if not value:
                    cc_coding = cc.get("coding", [{}])
                    value = cc_coding[0].get("display", "") if cc_coding else ""
            elif "valueBoolean" in resource:
                value = "Positive" if resource["valueBoolean"] else "Negative"

        if value is None or (isinstance(value, str) and not value.strip()):
            continue

        ref_range = ""
        ref_ranges = resource.get("referenceRange", [])
        if not isinstance(ref_ranges, list):
            ref_ranges = []
        if ref_ranges:
            rr = ref_ranges[0]
            if not isinstance(rr, dict):
                rr = {}
            low = rr.get("low", {}).get("value")
            high = rr.get("high", {}).get("value")
            rr_unit = rr.get("low", rr.get("high", {})).get("unit", unit)
            if low is not None and high is not None:
                ref_range = f"{low}-{high}"
                if rr_unit:
                    ref_range += f" {rr_unit}"
            elif low is not None:
                ref_range = f">={low}"
                if rr_unit:
                    ref_range += f" {rr_unit}"
            elif high is not None:
                ref_range = f"<={high}"
                if rr_unit:
                    ref_range += f" {rr_unit}"
            elif rr.get("text"):
                ref_range = rr["text"]

        specimen = ""
        spec_obj = resource.get("specimen", {})
        if not isinstance(spec_obj, dict):
            spec_obj = {}
        if spec_obj:
            specimen = (spec_obj.get("display", "") or "").strip()
            if not specimen:
                reference = (spec_obj.get("reference", "") or "").strip()
                if reference:
                    specimen = specimen_display_by_reference.get(reference, "")

        row_id = resource.get("id", "")
        if not row_id:
            identifiers = resource.get("identifier", [])
            row_id = identifiers[0].get("value", "") if identifiers else f"fhir_{index}"

        rows.append({
            "source_row_id": str(row_id),
            "source_lab_name": "",
            "source_panel_name": "",
            "source_test_name": test_name,
            "source_loinc": source_loinc,
            "raw_value": str(value),
            "source_unit": unit,
            "specimen_type": specimen,
            "source_reference_range": ref_range,
        })

    if not rows:
        raise ValueError("No Observation resources with numeric values found in FHIR input.")

    return rows


def _parse_hl7_sn(value: str) -> str:
    """Parse HL7 SN (Structured Numeric) value.

    Format: comparator^num1^separator^num2
    Examples: '<^10' -> '<10', '>^500' -> '>500', '^1^:^8' -> '1:8',
              '^100^-^200' -> '100-200', '^3^+' -> '3+'
    """
    parts = value.split("^")
    if len(parts) >= 4:
        comparator = parts[0].strip()
        num1 = parts[1].strip()
        separator = parts[2].strip()
        num2 = parts[3].strip()
        result = f"{comparator}{num1}{separator}{num2}".strip()
        return result if result else value
    if len(parts) == 3:
        comparator = parts[0].strip()
        num1 = parts[1].strip()
        suffix = parts[2].strip()
        result = f"{comparator}{num1}{suffix}".strip()
        return result if result else value
    if len(parts) == 2:
        comparator = parts[0].strip()
        number = parts[1].strip()
        if comparator and number:
            return f"{comparator}{number}"
        if number:
            return number
    return value


def read_hl7_input(path: Path) -> list[dict[str, str]]:
    """Read an HL7 v2.x file and extract OBX segments into input rows."""
    text = path.read_text(encoding="utf-8-sig")
    lines = [line.rstrip("\r") for line in text.split("\n") if line.strip()]

    # Detect field separator from MSH
    if not lines or not lines[0].startswith("MSH"):
        raise ValueError("Not a valid HL7 v2.x message: missing MSH segment.")

    field_sep = lines[0][3]  # Usually '|'
    comp_sep = lines[0][4] if len(lines[0]) > 4 else "^"  # Usually '^'

    rows: list[dict[str, str]] = []
    current_obr_name = ""
    current_specimen = ""
    row_id = 0

    for line in lines:
        fields = line.split(field_sep)
        segment = fields[0] if fields else ""

        if segment == "MSH":
            # Reset per-message state for batch files
            current_obr_name = ""
            current_specimen = ""

        if segment == "OBR" and len(fields) > 4:
            # OBR-4: Universal Service Identifier (test/panel name)
            obr4_parts = fields[4].split(comp_sep)
            current_obr_name = obr4_parts[1] if len(obr4_parts) > 1 else obr4_parts[0]
            # Reset specimen for new panel — prevents leakage from previous OBR
            current_specimen = ""
            # OBR-15: Specimen Source (component 1 = specimen type)
            if len(fields) > 15 and fields[15]:
                spec_parts = fields[15].split(comp_sep)
                current_specimen = spec_parts[0].strip()

        if segment == "SPM" and len(fields) > 4:
            # SPM-4: Specimen Type (HL7 v2.5+)
            spm4_parts = fields[4].split(comp_sep)
            current_specimen = spm4_parts[1] if len(spm4_parts) > 1 else spm4_parts[0]

        if segment == "OBX" and len(fields) > 5:
            value_type = fields[2] if len(fields) > 2 else ""  # NM, SN, ST, CE, etc.
            # OBX-3: Observation Identifier
            obx3_parts = fields[3].split(comp_sep)
            test_name = obx3_parts[1] if len(obx3_parts) > 1 else obx3_parts[0]
            source_loinc = ""
            if len(obx3_parts) > 2 and obx3_parts[2].strip().upper() in ("LN", "LOINC"):
                source_loinc = obx3_parts[0].strip()
            elif len(obx3_parts) > 5 and obx3_parts[5].strip().upper() in ("LN", "LOINC"):
                source_loinc = obx3_parts[3].strip()

            # OBX-5: Observation Value
            raw_value = fields[5] if len(fields) > 5 else ""
            if value_type == "SN":
                raw_value = _parse_hl7_sn(raw_value)

            # OBX-6: Units
            unit = ""
            if len(fields) > 6 and fields[6]:
                unit_parts = fields[6].split(comp_sep)
                unit = unit_parts[0]

            # OBX-7: Reference Range (don't append unit — normalizer uses
            # source_unit as fallback, and appending breaks numeric-prefix units
            # like 10*3/uL)
            ref_range = fields[7].strip() if len(fields) > 7 else ""

            # OBX-11: Result Status — skip cancelled/deleted/withdrawn results
            result_status = fields[11].strip().upper() if len(fields) > 11 else ""
            if result_status in ("X", "D", "W"):
                continue

            row_id += 1
            rows.append({
                "source_row_id": f"hl7_{row_id}",
                "source_lab_name": "",
                "source_panel_name": current_obr_name,
                "source_test_name": test_name,
                "source_loinc": source_loinc,
                "raw_value": raw_value,
                "source_unit": unit,
                "specimen_type": current_specimen,
                "source_reference_range": ref_range,
            })

    if not rows:
        raise ValueError("No OBX segments found in HL7 input.")

    return rows


_XSI = "{http://www.w3.org/2001/XMLSchema-instance}"
_HL7V3 = "{urn:hl7-org:v3}"
_LOINC_OID = "2.16.840.1.113883.6.1"


def _ccda_find(parent: ET.Element, tag: str) -> ET.Element | None:
    """Find a child element, trying without and with HL7v3 namespace."""
    el = parent.find(tag)
    if el is None:
        el = parent.find(f"{_HL7V3}{tag}")
    return el


def _ccda_original_text(parent: ET.Element | None) -> str:
    if parent is None:
        return ""
    original_text = _ccda_find(parent, "originalText")
    if original_text is None:
        return ""
    return (original_text.text or "").strip()


def _ccda_value_and_unit(element: ET.Element | None) -> tuple[str, str]:
    if element is None:
        return "", ""

    value = element.attrib.get("value", "").strip()
    unit = element.attrib.get("unit", "").strip()
    translation = _ccda_find(element, "translation")
    if translation is not None:
        if not value:
            value = translation.attrib.get("value", "").strip()
        if not unit:
            unit = translation.attrib.get("unit", "").strip() or _ccda_original_text(translation)
    return value, unit


def read_ccda_input(path: Path) -> list[dict[str, str]]:
    """Read a C-CDA XML document or fragment and extract lab observations."""
    content = path.read_text(encoding="utf-8-sig")

    # C-CDA examples are often fragments — wrap if needed
    if not content.strip().startswith("<?xml") and "<ClinicalDocument" not in content:
        content = (
            '<root xmlns:sdtc="urn:hl7-org:sdtc" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
            + content
            + "</root>"
        )

    root = _xml_fromstring(content)
    rows: list[dict[str, str]] = []
    row_id = 0

    # Search for observation elements with and without namespace
    observations = list(root.iter("observation")) + list(root.iter(f"{_HL7V3}observation"))

    for obs in observations:
        code_el = obs.find("code")
        if code_el is None:
            code_el = obs.find(f"{_HL7V3}code")
        value_el = obs.find("value")
        if value_el is None:
            value_el = obs.find(f"{_HL7V3}value")
        if code_el is None or value_el is None:
            continue

        # Get test name from code element
        loinc_code = ""
        if code_el.attrib.get("codeSystem") == _LOINC_OID:
            loinc_code = code_el.attrib.get("code", "")
        display_name = code_el.attrib.get("displayName", "")
        # Check translations for LOINC
        for trans in list(code_el.findall("translation")) + list(code_el.findall(f"{_HL7V3}translation")):
            if trans.attrib.get("codeSystem") == _LOINC_OID:
                if not display_name:
                    display_name = trans.attrib.get("displayName", "")
                if not loinc_code:
                    loinc_code = trans.attrib.get("code", "")

        test_name = display_name or loinc_code
        if not test_name:
            continue

        # Skip nullFlavor values that have no usable data
        nf = value_el.attrib.get("nullFlavor")
        if nf and not value_el.attrib.get("value") and _ccda_find(value_el, "translation") is None:
            continue

        # Get value based on xsi:type
        xsi_type = value_el.attrib.get(f"{_XSI}type", "")
        raw_value = ""
        unit = ""

        if xsi_type == "PQ":
            # Physical Quantity: value + unit
            raw_value, unit = _ccda_value_and_unit(value_el)
        elif xsi_type == "IVL_PQ":
            # Interval — used for "<10" style values
            low = _ccda_find(value_el, "low")
            high = _ccda_find(value_el, "high")
            low_value, low_unit = _ccda_value_and_unit(low)
            high_value, high_unit = _ccda_value_and_unit(high)
            if low is not None and low_value:
                raw_value = low_value
                unit = low_unit
                if low.attrib.get("inclusive") == "false":
                    raw_value = f">{raw_value}"
            elif high is not None and high_value:
                raw_value = high_value
                unit = high_unit
                if high.attrib.get("inclusive") == "false":
                    raw_value = f"<{raw_value}"
        elif xsi_type in ("INT", "REAL"):
            # Integer or real number
            raw_value = value_el.attrib.get("value", "").strip()
        elif xsi_type in ("ST", "ED"):
            # String or encapsulated data
            raw_value = (value_el.text or "").strip()
        elif xsi_type in ("CD", "CE", "CO"):
            # Coded value (qualitative)
            raw_value = value_el.attrib.get("displayName", "")
            if not raw_value:
                raw_value = value_el.attrib.get("code", "")

        if not raw_value:
            continue

        # Get reference range
        ref_range = ""
        ref_el = obs.find(".//referenceRange/observationRange/value")
        if ref_el is None:
            ref_el = obs.find(f".//{_HL7V3}referenceRange/{_HL7V3}observationRange/{_HL7V3}value")
        if ref_el is not None:
            ref_low = _ccda_find(ref_el, "low")
            ref_high = _ccda_find(ref_el, "high")
            low_val, low_unit = _ccda_value_and_unit(ref_low)
            high_val, high_unit = _ccda_value_and_unit(ref_high)
            ref_unit_str = low_unit or high_unit or unit
            if low_val and high_val:
                ref_range = f"{low_val}-{high_val}"
                if ref_unit_str:
                    ref_range += f" {ref_unit_str}"
            elif low_val:
                ref_range = f">={low_val}"
                if ref_unit_str:
                    ref_range += f" {ref_unit_str}"
            elif high_val:
                ref_range = f"<={high_val}"
                if ref_unit_str:
                    ref_range += f" {ref_unit_str}"

        # Get specimen type from specimen/specimenRole/specimenPlayingEntity/code
        specimen = ""
        for spec_path in (
            ".//specimen/specimenRole/specimenPlayingEntity/code",
            f".//{_HL7V3}specimen/{_HL7V3}specimenRole/{_HL7V3}specimenPlayingEntity/{_HL7V3}code",
        ):
            spec_el = obs.find(spec_path)
            if spec_el is not None:
                specimen = spec_el.attrib.get("displayName", "")
                if not specimen:
                    specimen = spec_el.attrib.get("code", "")
                break

        row_id += 1
        rows.append({
            "source_row_id": f"ccda_{row_id}",
            "source_lab_name": "",
            "source_panel_name": "",
            "source_test_name": test_name,
            "source_loinc": loinc_code,
            "raw_value": raw_value,
            "source_unit": unit,
            "specimen_type": specimen,
            "source_reference_range": ref_range,
        })

    if not rows:
        raise ValueError(
            "No completed observation values found in C-CDA input. "
            "The document may contain only pending or null-flavored results."
        )

    return rows


def read_excel_input(path: Path) -> list[dict[str, str]]:
    """Read an Excel file (.xlsx/.xls) and extract rows using the header row for column names.

    Attempts to match columns to our required schema by flexible header matching.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel ingest: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        if len(sheet_names) > 1:
            logger.warning(
                "Excel file '%s' has %d sheets (%s). Only the active sheet "
                "will be read; data in other sheets will be ignored.",
                path.name,
                len(sheet_names),
                ", ".join(repr(s) for s in sheet_names),
            )

        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no active worksheet.")

        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            raise ValueError("Excel file has no header row.")

        # Normalize headers: strip, lowercase, replace spaces/dashes with underscores
        def norm_header(h: str) -> str:
            return h.strip().lower().replace(" ", "_").replace("-", "_")

        headers = [norm_header(str(h or "")) for h in raw_headers]

        # Map flexible header names to our schema
        header_map = {
            "source_row_id": {"source_row_id", "row_id", "id", "sample_id", "accession", "accession_number"},
            "source_test_name": {"source_test_name", "test_name", "test", "analyte", "component", "observation", "lab_test", "result_name"},
            "raw_value": {"raw_value", "value", "result", "result_value", "observed_value"},
            "source_unit": {"source_unit", "unit", "units", "uom", "unit_of_measure"},
            "specimen_type": {"specimen_type", "specimen", "sample_type", "fluid", "matrix"},
            "source_reference_range": {"source_reference_range", "reference_range", "ref_range", "normal_range", "range"},
            "source_lab_name": {"source_lab_name", "lab_name", "lab", "laboratory", "performing_lab"},
            "source_panel_name": {"source_panel_name", "panel_name", "panel", "order_name", "test_group"},
            "source_loinc": {"source_loinc", "loinc", "loinc_code", "observation_loinc"},
        }

        col_map: dict[int, str] = {}
        for idx, header in enumerate(headers):
            for canonical, variants in header_map.items():
                if header in variants:
                    col_map[idx] = canonical
                    break

        duplicate_mapped = _find_duplicate_labels(list(col_map.values()))
        if duplicate_mapped:
            raise ValueError(
                "Excel file has duplicate columns mapping to the same field: "
                f"{', '.join(duplicate_mapped)}"
            )

        # Verify required columns found
        mapped_cols = set(col_map.values())
        required = {"source_test_name", "raw_value"}
        missing = required - mapped_cols
        if missing:
            raise ValueError(
                f"Excel file missing required columns: {', '.join(missing)}. "
                f"Found headers: {[str(h) for h in raw_headers]}"
            )

        rows: list[dict[str, str]] = []
        row_num = 0
        for data_row in rows_iter:
            row_num += 1
            row: dict[str, str] = {
                "source_row_id": str(row_num),
                "source_lab_name": "",
                "source_panel_name": "",
                "source_test_name": "",
                "source_loinc": "",
                "raw_value": "",
                "source_unit": "",
                "specimen_type": "",
                "source_reference_range": "",
            }
            for idx, val in enumerate(data_row):
                if idx in col_map:
                    row[col_map[idx]] = str(val).strip() if val is not None else ""

            # Skip completely empty rows
            if not row["source_test_name"] and not row["raw_value"]:
                continue

            # Auto-generate row ID if not mapped
            if "source_row_id" not in mapped_cols:
                row["source_row_id"] = str(row_num)

            rows.append(row)
    finally:
        wb.close()

    if not rows:
        raise ValueError("Excel file has no data rows.")

    return rows


def _detect_csv_dialect(path: Path) -> csv.Dialect | None:
    """Auto-detect CSV delimiter by sniffing the first line."""
    try:
        with path.open("r", encoding="utf-8-sig") as f:
            sample = f.read(4096)
    except UnicodeDecodeError:
        with path.open("r", encoding="latin-1") as f:
            sample = f.read(4096)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return None


def read_input_csv(path: Path) -> list[dict[str, str]]:
    dialect = _detect_csv_dialect(path)
    try:
        handle = path.open("r", encoding="utf-8-sig", newline="")
        handle.read(1)  # Test read to detect encoding errors early
        handle.seek(0)
    except UnicodeDecodeError:
        handle.close()  # Close the failed UTF-8 handle before opening Latin-1
        handle = path.open("r", encoding="latin-1", newline="")
    with handle:
        reader = csv.DictReader(handle, dialect=dialect) if dialect else csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError("Input CSV has no header row.")

        duplicate_headers = _find_duplicate_labels([str(field) for field in reader.fieldnames])
        if duplicate_headers:
            raise ValueError(f"Input CSV has duplicate columns: {', '.join(duplicate_headers)}")

        missing = [column for column in REQUIRED_INPUT_COLUMNS if column not in reader.fieldnames]
        if missing:
            raise ValueError(f"Input CSV is missing required columns: {', '.join(missing)}")

        rows = [{key: (value or "") for key, value in row.items() if key is not None} for row in reader]
        if not rows:
            raise ValueError("Input CSV has a header row but no data rows.")
        return rows


def write_result(result: NormalizationResult, output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "normalized_records.json"
    csv_path = output_dir / "normalized_records.csv"

    json_path.write_text(
        json.dumps(result.to_json_dict(), indent=2) + "\n",
        encoding="utf-8",
    )

    fieldnames = list(result.records[0].to_csv_row().keys()) if result.records else [
        "source_row_number",
        "source_row_id",
        "source_lab_name",
        "source_panel_name",
        "source_test_name",
        "alias_key",
        "raw_value",
        "source_unit",
        "specimen_type",
        "source_reference_range",
        "canonical_biomarker_id",
        "canonical_biomarker_name",
        "loinc",
        "mapping_status",
        "match_confidence",
        "status_reason",
        "mapping_rule",
        "normalized_value",
        "normalized_unit",
        "normalized_reference_range",
        "provenance_source_row_id",
        "provenance_alias_key",
    ]

    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in result.records:
            writer.writerow(record.to_csv_row())

    return json_path, csv_path


def write_fhir_bundle(result: NormalizationResult, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    fhir_path = output_dir / "fhir_observations.json"
    fhir_path.write_text(
        json.dumps(build_bundle(result), indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    return fhir_path


def write_summary_report(result: NormalizationResult, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "normalization_summary.md"
    report_path.write_text(build_summary_report(result), encoding="utf-8")
    return report_path
